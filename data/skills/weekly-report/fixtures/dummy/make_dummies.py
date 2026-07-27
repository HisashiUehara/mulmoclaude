#!/usr/bin/env python3
"""Stage 2 用ダミー入力の生成 (自作)。

- contractor-report.xlsx : ダミー業者レポート(日本語)。原文は仮ゴールド expected.dsl.json と
  同一セル・同一文にして、抽出エンジンの正しさをゴールドで検証できるようにする。
- tenant-template.xlsx    : ダミーのテナント用テンプレート(英語)。固定ラベル + {{NNN}} プレースホルダ
  (NNN は DSL の項目ID)。ステップ6の流し込み先。書式(フォント/サイズ/配置)付き。

NDA未確認のため実データは使わない。ここで作るのは全て架空のダミー。
"""
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).resolve().parent
GOLD = json.loads((HERE.parent / "gold" / "expected.dsl.json").read_text(encoding="utf-8"))


def build_contractor() -> None:
    """仮ゴールドの原文を、その source_location のセルに配置した業者レポート。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "週報"
    meiryo = "Meiryo"
    for it in GOLD["items"]:
        sheet, coord = it["source_location"].split("/", 1)  # 例: 週報/B2
        cell = ws[coord]
        cell.value = it["原文"]
        # 先頭行(タイトル)は太字大きめ、他は本文。書式抽出の検証材料にする。
        if it["id"] == "001":
            cell.font = Font(name=meiryo, size=14, bold=True)
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.font = Font(name=meiryo, size=11)
    ws.column_dimensions["B"].width = 60
    wb.save(HERE / "contractor-report.xlsx")


# テナントテンプレートの構成: (セル, 種別, テキスト)
#   種別 heading=章見出し, label=項目ラベル, ph=プレースホルダ({{ID}}), th=表ヘッダ
_TEMPLATE = [
    ("A1", "ph", "{{001}}"),
    ("A3", "label", "Project:"), ("B3", "ph", "{{002}}"),
    ("A4", "label", "Reporting week:"), ("B4", "ph", "{{003}}"),
    ("A5", "label", "Contractor:"), ("B5", "ph", "{{004}}"),
    ("A7", "heading", "1. Overall Progress"),
    ("A8", "ph", "{{005}}"),
    ("A10", "heading", "2. Progress by Area"),
    ("A11", "th", "Area / This Week"),
    ("A12", "ph", "{{006}}"), ("A13", "ph", "{{007}}"), ("A14", "ph", "{{008}}"),
    ("A16", "heading", "3. Plan for Next Week"), ("A17", "ph", "{{009}}"),
    ("A19", "heading", "4. Issues"), ("A20", "ph", "{{010}}"),
    ("A22", "heading", "5. Safety"), ("A23", "ph", "{{011}}"),
    ("A25", "heading", "6. Notes"), ("A26", "ph", "{{012}}"),
]


def _style(cell, kind: str) -> None:
    arial = "Arial"
    if kind == "ph" and cell.coordinate == "A1":  # タイトルプレースホルダ
        cell.font = Font(name=arial, size=16, bold=True)
        cell.alignment = Alignment(horizontal="center")
    elif kind == "heading":
        cell.font = Font(name=arial, size=12, bold=True)
    elif kind == "label":
        cell.font = Font(name=arial, size=11, bold=True)
    elif kind == "th":
        cell.font = Font(name=arial, size=11, bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
    else:  # ph 本文
        cell.font = Font(name=arial, size=11)


def build_template() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.merge_cells("A1:C1")
    for coord, kind, text in _TEMPLATE:
        ws[coord].value = text
        _style(ws[coord], kind)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 40
    wb.save(HERE / "tenant-template.xlsx")


if __name__ == "__main__":
    build_contractor()
    build_template()
    print("wrote contractor-report.xlsx, tenant-template.xlsx")
